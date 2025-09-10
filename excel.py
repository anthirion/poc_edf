from openpyxl import load_workbook

capitals_data = {
    "France": "Paris",
    "Canada": "Ottawa",
    "Brazil": "Brasilia",
}

regions_data = {
    "France": "Europe",
    "Canada": "North America",
    "Brazil": "South America",
}


def get_column_indexes(sheet):
  headers = [cell.value for cell in sheet[1]]
  try:
    country_col_idx = headers.index("Country") + 1
    capital_col_idx = headers.index("Capital city") + 1
    region_col_idx = headers.index("Region") + 1
    return country_col_idx, capital_col_idx, region_col_idx
  except ValueError as e:
    print(f"Error: Missing column in Excel file - {e}")
    return


def update_columns(sheet, country_col_idx, columns_to_update):
  for row in range(2, sheet.max_row + 1):
    country_cell = sheet.cell(row=row, column=country_col_idx)
    country_name = country_cell.value

    if country_name in capitals_data:
      capital_cell = sheet.cell(row=row, column=columns_to_update[0])
      capital_cell.value = capitals_data[country_name]
      region_cell = sheet.cell(row=row, column=columns_to_update[1])
      region_cell.value = regions_data[country_name]


def fill_excel(filename, capitals_data=None, regions_data=None):
  """
  Updates the capital cities in an Excel file based on a dictionary of countries and capitals.
  """
  if capitals_data is None:
    capitals_data = {}

  if regions_data is None:
    regions_data = {}

  try:
    workbook = load_workbook(filename)
    sheet = workbook.active

    country_col_idx, *columns_to_update = get_column_indexes(sheet)
    update_columns(sheet, country_col_idx, columns_to_update)

    workbook.save(filename)
    print(f"Successfully saved changes to {filename}")

  except FileNotFoundError:
    print(f"Error: The file {filename} was not found.")
  except Exception as e:
    print(f"An unexpected error occurred: {e}")


if __name__ == "__main__":
  fill_excel("countries_info.xlsx", capitals_data, regions_data)
